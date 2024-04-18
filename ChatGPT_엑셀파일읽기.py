from openpyxl import load_workbook

# Excel 파일 읽기
def read_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    # 각 행마다 데이터 출력
    for row in ws.iter_rows(values_only=True):
        print(row)

# 생성된 Excel 파일 읽기
read_excel("c:\\work\\products.xlsx")
